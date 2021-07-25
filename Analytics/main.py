# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
import openpyxl


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.

def make_file(filename,file_path):
    if os.path.exists(os.path.join(file_path, filename)):
        print(f'{filename} already exists')
    else:
        f = open(os.path.join(file_path,filename),'w')
        f.close()
        # with open(os.path.join(file_path, filename), 'w', encoding='utf-8') as f:
        #     print(f'{filename} created')

def create_xl_workbook(filename,file_path):
    xl_name = os.path.join(file_path, filename)
    if os.path.exists(xl_name):
        print(f'{filename} already exists')
    else:
        # file = os.path.join(file_path, filename)
        wb = openpyxl.Workbook()
        wb.create_sheet(title='january_2013', index=0)
        wb.create_sheet(title='february_2013', index=1)
        wb.create_sheet(title='march_2013', index=2)
        wb.save(xl_name)
    get_sheetnames(xl_name)
    # remove_sheet(xl_name='sheet')

def get_sheetnames(xl_name):
    wb = openpyxl.load_workbook(filename=xl_name)
    print(f'sheetnames:{wb.sheetnames}')
    if 'Sheet' in wb.sheetnames:
        remove_sheet(xl_name, sheet_name='Sheet')
    else:
        print('no worksheet named sheet')

def remove_sheet(xl_name, sheet_name):
    wb = openpyxl.load_workbook(filename=xl_name)
    ws = wb['Sheet']
    print(f'{wb.remove_sheet(ws)} is removed')
    wb.save(xl_name)



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')
    make_file('1excel_introspect_workbook.py', os.getcwd())
    create_xl_workbook('sales_2013.xlsx', os.getcwd())

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
